from distutils.command.clean import clean
from locale import currency
from textwrap import indent
import pandas as pd
from openpyxl import load_workbook
import os
import shutil


def get_sheetnames_xlsx(filepath):
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    return wb.sheetnames

def clean_dataframe(df):

    for column in df.columns:
        df_cleaned = df
        df_cleaned[column] = df[column].apply(lambda x: x.split(" ")[0] if "ms" in x else x)
    
    return df_cleaned

if __name__ == '__main__':
    
    if os.path.exists('../measures'):
        shutil.rmtree("../measures")
   
    # creamos el directorio donde se almacenaran las medidas
    os.makedirs('../measures')

    sheets_list = get_sheetnames_xlsx("measures.xlsx")
    for sheet in sheets_list:
        current_df = pd.read_excel("measures.xlsx", sheet)
        cleaned_df = clean_dataframe(current_df)
        cleaned_df.to_csv("../measures/"+sheet+".csv", index=False)
