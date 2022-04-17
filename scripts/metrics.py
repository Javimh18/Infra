from distutils.command.clean import clean
from locale import currency
from textwrap import indent
import pandas as pd
from openpyxl import load_workbook
import numpy as np

def get_sheetnames_xlsx(filepath):
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    return wb.sheetnames

if __name__ == '__main__':
    df_unican = pd.read_csv("../rtt_unican.csv")
    print("La varianza calculada para Unican: ", np.var(df_unican['RTT(ms)']))

    for sheet in get_sheetnames_xlsx("../measures.xlsx"):
        path = "../measures/"+sheet+".csv"
        df_current = pd.read_csv(path)
        print("La varianza para ", sheet, ":", np.var(df_current['Mean']))
    

    